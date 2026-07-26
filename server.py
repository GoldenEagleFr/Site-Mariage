from __future__ import annotations

import email.mime.multipart
import email.mime.text
import hashlib
import json
import os
import smtplib
import threading
import time
import unicodedata
from collections import defaultdict
from datetime import datetime
from html import escape
from functools import partial
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR = Path(__file__).resolve().parent
ACCESS_CODE = ""
ACCESS_COOKIE_NAME = "mariage_access"
DATA_FILE = BASE_DIR / "data.json"
BACKUP_FILE = BASE_DIR / "data.backup.json"
BACKUPS_DIR = BASE_DIR / "backups"
BUDGET_EXCEL_FILE = BASE_DIR / "budget_mariage.xlsx"
DATA_LOCK = threading.Lock()
ADMIN_PASSWORD = ""
DEFAULT_ADMIN_PASSWORD = "mariage2026"
_BACKUP_TIMER: threading.Timer | None = None
BACKUP_INTERVAL_SECONDS = 6 * 3600  # 6 heures
BACKUP_MAX_COUNT = 10
MAINTENANCE_FILE = BASE_DIR / "maintenance.html"

_RATE_LIMIT: defaultdict = defaultdict(list)
_RATE_LIMIT_LOCK = threading.Lock()


# ── Backup périodique ───────────────────────────────────────────

def run_periodic_backup() -> None:
    """Copie data.json dans backups/ et supprime les anciens backups."""
    global _BACKUP_TIMER
    try:
        if DATA_FILE.exists():
            BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            dest = BACKUPS_DIR / f"data.backup.{stamp}.json"
            dest.write_bytes(DATA_FILE.read_bytes())
            # Garder max BACKUP_MAX_COUNT fichiers (les plus récents)
            existing = sorted(BACKUPS_DIR.glob("data.backup.*.json"), key=lambda p: p.name)
            for old in existing[:-BACKUP_MAX_COUNT]:
                try:
                    old.unlink()
                except OSError:
                    pass
    except OSError as exc:
        print(f"Avertissement backup : {exc}", flush=True)
    finally:
        # Replanifier
        _BACKUP_TIMER = threading.Timer(BACKUP_INTERVAL_SECONDS, run_periodic_backup)
        _BACKUP_TIMER.daemon = True
        _BACKUP_TIMER.start()


def start_backup_scheduler() -> None:
    global _BACKUP_TIMER
    _BACKUP_TIMER = threading.Timer(BACKUP_INTERVAL_SECONDS, run_periodic_backup)
    _BACKUP_TIMER.daemon = True
    _BACKUP_TIMER.start()
    print(f"Backup automatique activé (toutes les {BACKUP_INTERVAL_SECONDS // 3600}h, max {BACKUP_MAX_COUNT} fichiers)", flush=True)


# ── Notification email RSVP ────────────────────────────────────

def send_rsvp_notification(guest: dict) -> None:
    """Envoie un email de notification aux mariés après un RSVP. Si les variables
    d'environnement ne sont pas définies, la fonction retourne silencieusement."""
    smtp_host = os.getenv("MARIAGE_SMTP_HOST", "").strip()
    smtp_user = os.getenv("MARIAGE_SMTP_USER", "").strip()
    smtp_pass = os.getenv("MARIAGE_SMTP_PASSWORD", "").strip()
    notify_email = os.getenv("MARIAGE_NOTIFY_EMAIL", "").strip()
    if not all([smtp_host, smtp_user, smtp_pass, notify_email]):
        return

    smtp_port = int(os.getenv("MARIAGE_SMTP_PORT", "587"))
    name      = guest.get("name", "Inconnu")
    status    = "Oui" if guest.get("status") == "yes" else "Non"
    allergies = guest.get("allergies", "") or "—"
    heberg    = "Oui" if guest.get("hebergement") else "Non"

    subject = f"RSVP reçu — {name}"
    body = (
        f"Nouveau RSVP reçu\n\n"
        f"Invité   : {name}\n"
        f"Réponse  : {status}\n"
        f"Allergies : {allergies}\n"
        f"Hébergement : {heberg}\n"
    )

    msg = email.mime.multipart.MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = notify_email
    msg["Subject"] = subject
    msg.attach(email.mime.text.MIMEText(body, "plain", "utf-8"))

    def _send() -> None:
        try:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
                server.ehlo()
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, notify_email, msg.as_string())
        except Exception as exc:
            print(f"Avertissement email RSVP : {exc}", flush=True)

    t = threading.Thread(target=_send, daemon=True)
    t.start()


_404_PAGE = """\
<!doctype html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Page introuvable — Florence &amp; Antoine</title>
  <style>
    body{font-family:Manrope,sans-serif;background:#faf9f7;color:#333;display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:100vh;margin:0;text-align:center;padding:2rem;box-sizing:border-box}
    h1{font-size:4rem;margin:0 0 .5rem}
    p{margin:.4rem 0;color:#666}
    a{color:#c0392b;font-weight:600;text-decoration:none}
    a:hover{text-decoration:underline}
  </style>
</head>
<body>
  <h1>\U0001f48d</h1>
  <p style="font-size:1.4rem;font-weight:700;color:#333">Page introuvable</p>
  <p>Cette page n’existe pas ou a été déplacée.</p>
  <p style="margin-top:1rem"><a href="/">← Retour au site</a></p>
</body>
</html>"""

DEFAULT_DATA = {
    "budgetGoal": 15000,
    "budgetGuestCount": 150,
    "budgetAdultCount": 110,
    "budgetItems": [],
    "budgetCategories": [],
    "tasks": [],
    "guests": [],
    "guestGroups": [],
    "updatedAt": 0,
}

VALID_GUEST_STATUS = {"pending", "yes", "no"}
VALID_GUEST_GROUP_TYPE = {"single", "couple", "family"}
VALID_GUEST_ATTENDANCE_TYPE = {"vin_repas", "vin_only"}


def _compute_access_token(code: str) -> str:
    return hashlib.sha256(f"mariage_v1_{code}".encode()).hexdigest()[:48]


def normalize_guest_group_type(value: object) -> str:
    group_type = str(value or "").strip()
    if group_type in VALID_GUEST_GROUP_TYPE:
        return group_type
    return "single"


def normalize_guest_attendance_type(value: object) -> str:
    attendance_type = str(value or "").strip()
    if attendance_type in VALID_GUEST_ATTENDANCE_TYPE:
        return attendance_type
    return "vin_repas"


def normalize_guest_party_size(group_type: str, value: object) -> int:
    if group_type == "single":
        return 1
    if group_type == "couple":
        return 2

    if isinstance(value, (int, float)):
        size = int(value)
    else:
        try:
            size = int(str(value).strip())
        except (TypeError, ValueError):
            size = 0

    if size < 1:
        return 3
    return size


def create_default_data() -> dict:
    return {
        "budgetGoal": DEFAULT_DATA["budgetGoal"],
        "budgetGuestCount": DEFAULT_DATA["budgetGuestCount"],
        "budgetAdultCount": DEFAULT_DATA["budgetAdultCount"],
        "budgetItems": [],
        "budgetCategories": [],
        "tasks": [],
        "guests": [],
        "guestGroups": [],
        "updatedAt": DEFAULT_DATA["updatedAt"],
    }


def normalize_data(candidate: object) -> dict:
    if not isinstance(candidate, dict):
        return create_default_data()

    normalized = create_default_data()

    budget_goal = candidate.get("budgetGoal")
    if isinstance(budget_goal, (int, float)) and budget_goal >= 0:
        normalized["budgetGoal"] = int(budget_goal)

    budget_guest_count = candidate.get("budgetGuestCount")
    if isinstance(budget_guest_count, (int, float)) and int(budget_guest_count) >= 1:
        normalized["budgetGuestCount"] = int(budget_guest_count)

    budget_adult_count = candidate.get("budgetAdultCount")
    if isinstance(budget_adult_count, (int, float)) and int(budget_adult_count) >= 1:
        normalized["budgetAdultCount"] = int(budget_adult_count)

    normalized["budgetAdultCount"] = min(normalized["budgetAdultCount"], normalized["budgetGuestCount"])

    updated_at = candidate.get("updatedAt")
    if isinstance(updated_at, (int, float)) and updated_at >= 0:
        normalized["updatedAt"] = int(updated_at)

    budget_items = candidate.get("budgetItems")
    if isinstance(budget_items, list):
        cleaned_budget = []
        for item in budget_items:
            if not isinstance(item, dict):
                continue
            label = str(item.get("label", "")).strip()
            if not label:
                continue
            identifier = str(item.get("id", "")).strip() or "item"
            category_id = str(item.get("categoryId", "")).strip() or None
            sup = item.get("supplier", {})
            sup = sup if isinstance(sup, dict) else {}

            # Versements partiels (optionnel)
            raw_payments = item.get("payments")
            cleaned_payments = []
            if isinstance(raw_payments, list):
                for pmt in raw_payments:
                    if not isinstance(pmt, dict):
                        continue
                    pmt_label = str(pmt.get("label", "")).strip() or "Versement"
                    pmt_total_raw = pmt.get("amountTotal", 0)
                    pmt_paid_raw  = pmt.get("amountPaid", 0)
                    pmt_total = float(pmt_total_raw) if isinstance(pmt_total_raw, (int, float)) else 0.0
                    pmt_paid  = float(pmt_paid_raw)  if isinstance(pmt_paid_raw,  (int, float)) else 0.0
                    cleaned_payments.append({
                        "id":          str(pmt.get("id", "")).strip() or "pmt",
                        "label":       pmt_label,
                        "amountTotal": max(0.0, pmt_total),
                        "amountPaid":  max(0.0, pmt_paid),
                        "dueDate":     str(pmt.get("dueDate", "")).strip(),
                        "payer":       str(pmt.get("payer", "")).strip(),
                        "solde":       bool(pmt.get("solde", False)),
                    })

            # Totaux calculés depuis les versements si présents, sinon flat
            if cleaned_payments:
                amount_total = sum(p["amountTotal"] for p in cleaned_payments)
                amount_paid  = sum(p["amountPaid"]  for p in cleaned_payments)
            else:
                amount_total_raw = item.get("amountTotal", item.get("amount", 0))
                amount_paid_raw  = item.get("amountPaid", 0)
                amount_total = float(amount_total_raw) if isinstance(amount_total_raw, (int, float)) else 0.0
                amount_paid  = float(amount_paid_raw)  if isinstance(amount_paid_raw,  (int, float)) else 0.0

            entry = {
                "id": identifier,
                "label": label,
                "categoryId": category_id,
                "amountTotal": max(0.0, amount_total),
                "amountPaid":  max(0.0, amount_paid),
                "solde": bool(item.get("solde", False)),
                "dueDate": str(item.get("dueDate", "")).strip(),
                "supplier": {
                    "contact": str(sup.get("contact", "")).strip(),
                    "phone":   str(sup.get("phone",   "")).strip(),
                    "email":   str(sup.get("email",   "")).strip(),
                    "address": str(sup.get("address", "")).strip(),
                    "website": str(sup.get("website", "")).strip(),
                    "notes":   str(sup.get("notes",   "")).strip(),
                },
            }
            if cleaned_payments:
                entry["payments"] = cleaned_payments
            cleaned_budget.append(entry)
        normalized["budgetItems"] = cleaned_budget

    budget_categories = candidate.get("budgetCategories")
    if isinstance(budget_categories, list):
        cleaned_categories = []
        for cat in budget_categories:
            if not isinstance(cat, dict):
                continue
            name = str(cat.get("name", "")).strip()
            if not name:
                continue
            cat_id = str(cat.get("id", "")).strip() or "cat"
            cleaned_categories.append({"id": cat_id, "name": name})
        normalized["budgetCategories"] = cleaned_categories

    tasks = candidate.get("tasks")
    if isinstance(tasks, list):
        cleaned_tasks = []
        for task in tasks:
            if not isinstance(task, dict):
                continue
            text = str(task.get("text", "")).strip()
            if not text:
                continue
            identifier = str(task.get("id", "")).strip() or "task"
            cleaned_tasks.append(
                {
                    "id": identifier,
                    "text": text,
                    "done": bool(task.get("done", False)),
                }
            )
        normalized["tasks"] = cleaned_tasks

    guests = candidate.get("guests")
    if isinstance(guests, list):
        cleaned_guests = []
        for guest in guests:
            if not isinstance(guest, dict):
                continue
            name = str(guest.get("name", "")).strip()
            status = guest.get("status", "pending")
            group_type = normalize_guest_group_type(guest.get("groupType"))
            attendance_type = normalize_guest_attendance_type(guest.get("attendanceType"))
            if not name:
                continue
            cleaned_guests.append(
                {
                    "id": str(guest.get("id", "")).strip() or "guest",
                    "name": name,
                    "groupType": group_type,
                    "attendanceType": attendance_type,
                    "partySize": normalize_guest_party_size(group_type, guest.get("partySize")),
                    "isHost": bool(guest.get("isHost", False)),
                    "status": "yes" if bool(guest.get("isHost", False)) else (status if status in VALID_GUEST_STATUS else "pending"),
                    "rsvpToken": str(guest.get("rsvpToken", "")).strip(),
                    "hebergement": bool(guest.get("hebergement", False)),
                    "hebergementInfo": bool(guest.get("hebergementInfo", False)),
                    "rsvpSubmittedAt": int(guest.get("rsvpSubmittedAt", 0) or 0),
                    "guestCategory": str(guest.get("guestCategory", "")).strip() if str(guest.get("guestCategory", "")).strip() in {"adult", "teen", "child"} else "adult",
                    "musicSuggestion": str(guest.get("musicSuggestion", "")).strip(),
                    "allergies": str(guest.get("allergies", "")).strip(),
                    "otherQuestion": str(guest.get("otherQuestion", "")).strip(),
                    "colorGroupId": (lambda v: str(v).strip() or None if isinstance(v, str) else None)(guest.get("colorGroupId")),
                }
            )
        normalized["guests"] = cleaned_guests

    guest_groups = candidate.get("guestGroups")
    if isinstance(guest_groups, list):
        cleaned_groups = []
        for grp in guest_groups:
            if not isinstance(grp, dict):
                continue
            grp_name = str(grp.get("name", "")).strip()
            if not grp_name:
                continue
            grp_color = str(grp.get("color", "#888888")).strip()
            if not grp_color.startswith("#"):
                grp_color = "#888888"
            cleaned_groups.append({
                "id":    str(grp.get("id", "")).strip() or "grp",
                "name":  grp_name,
                "color": grp_color,
            })
        normalized["guestGroups"] = cleaned_groups

    return normalized


def write_data_file(data: dict) -> None:
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    if DATA_FILE.exists():
        try:
            BACKUP_FILE.write_bytes(DATA_FILE.read_bytes())
        except OSError:
            pass
    DATA_FILE.write_text(payload, encoding="utf-8")


def write_budget_excel(data: dict) -> None:  # noqa: PLR0912, PLR0915
    budget_items = data.get("budgetItems", [])
    guests       = data.get("guests", [])
    budget_goal  = float(data.get("budgetGoal", 0) or 0)
    guest_count  = int(data.get("budgetGuestCount", 0) or 0)
    adult_count  = int(data.get("budgetAdultCount", 0) or 0)
    updated_at   = int(data.get("updatedAt", 0) or 0)
    categories_by_id = {c.get("id"): c.get("name", "") for c in data.get("budgetCategories", [])}

    total_prevu = sum(float(i.get("amountTotal", i.get("amount", 0)) or 0) for i in budget_items)
    total_paye  = sum(float(i.get("amountPaid", 0) or 0) for i in budget_items)
    total_reste = max(total_prevu - total_paye, 0)
    budget_libre = max(budget_goal - total_prevu, 0)
    over_budget  = max(total_prevu - budget_goal, 0)
    cost_per_guest = total_prevu / guest_count if guest_count > 0 else 0
    cost_per_adult = total_prevu / adult_count if adult_count > 0 else 0

    by_category: dict = defaultdict(list)
    for item in budget_items:
        by_category[item.get("categoryId") or "__none__"].append(item)

    def cat_sort_key(cat_id: str) -> str:
        return "ZZZZ" if cat_id == "__none__" else categories_by_id.get(cat_id, "ZZZZ")

    # ── Palette ──────────────────────────────────────────────────
    PINK_DARK  = "8A2C4F"; PINK = "F03F59"; PINK_LT = "F6DCE3"; PINK_XLT = "FCEEF3"
    GREEN      = "1E8449"; GREEN_LT  = "D5F5E3"
    ORANGE     = "BA4A00"; ORANGE_LT = "FDEBD0"
    BLUE       = "1F618D"; BLUE_LT   = "D6EAF8"
    RED        = "C0392B"; RED_LT    = "FADBD8"
    GREY       = "7F8C8D"; GREY_LT   = "F2F3F4"
    WHITE      = "FFFFFF"; BLACK     = "1C1C1C"

    # ── Helpers ──────────────────────────────────────────────────
    def fl(c: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=c)

    def fn(bold: bool = False, color: str = BLACK, size: int = 11, italic: bool = False) -> Font:
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    def bd(color: str = "D5D8DC") -> Border:
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    AL_C = Alignment(horizontal="center", vertical="center")
    AL_L = Alignment(horizontal="left",   vertical="center")
    AL_R = Alignment(horizontal="right",  vertical="center")
    AL_W = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    EUR = '#,##0.00 "€"'; PCT = "0%"

    def sc(cell, fnt=None, fll=None, brd=None, aln=None, fmt=None) -> None:
        if fnt: cell.font      = fnt
        if fll: cell.fill      = fll
        if brd: cell.border    = brd
        if aln: cell.alignment = aln
        if fmt: cell.number_format = fmt

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # ONGLET 1 — Synthèse
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Synthèse"
    ws1.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [2, 24, 18, 18, 18, 18, 2, 2]):
        ws1.column_dimensions[col].width = w

    # Titre
    ws1.merge_cells("B2:F2")
    sc(ws1["B2"], fnt=fn(bold=True, color=PINK_DARK, size=20),
       fll=fl(PINK_LT), brd=bd(PINK), aln=AL_C)
    ws1["B2"].value = "Budget Mariage"
    ws1.row_dimensions[2].height = 42

    ws1.merge_cells("B3:F3")
    c3 = ws1["B3"]
    if updated_at > 0:
        c3.value = datetime.fromtimestamp(updated_at / 1000)
        c3.number_format = '"Mis à jour le "DD/MM/YYYY" à "HH:mm'
    else:
        c3.value = "—"
    sc(c3, fnt=fn(italic=True, color=GREY, size=9), fll=fl(PINK_XLT), aln=AL_C)
    ws1.row_dimensions[3].height = 14

    # ── 3 cartes métriques (lignes 5-8) ─────────────────────────
    CARDS = [
        ("Montant prévu",  total_prevu, BLUE,   BLUE_LT,   "Total des devis engagés"),
        ("Déjà payé",      total_paye,  GREEN,  GREEN_LT,  "Acomptes et paiements versés"),
        ("Reste à régler", total_reste, ORANGE if total_reste > 0 else GREY,
                                        ORANGE_LT if total_reste > 0 else GREY_LT,
                                        "Sur devis déjà engagés"),
    ]
    CR = 5
    for i, (lbl, val, col, bg, sub) in enumerate(CARDS):
        c0 = 2 + i * 2
        ws1.merge_cells(start_row=CR,   start_column=c0, end_row=CR,   end_column=c0+1)
        ws1.merge_cells(start_row=CR+1, start_column=c0, end_row=CR+2, end_column=c0+1)
        ws1.merge_cells(start_row=CR+3, start_column=c0, end_row=CR+3, end_column=c0+1)
        hc = ws1.cell(row=CR,   column=c0, value=lbl)
        vc = ws1.cell(row=CR+1, column=c0, value=val)
        sc2 = ws1.cell(row=CR+3, column=c0, value=sub)
        sc(hc,  fnt=fn(bold=True, color=WHITE, size=10), fll=fl(col),  brd=bd(col), aln=AL_C)
        sc(vc,  fnt=fn(bold=True, color=col, size=18),   fll=fl(bg),   brd=bd(col), aln=AL_C, fmt=EUR)
        sc(sc2, fnt=fn(italic=True, color=col, size=9),  fll=fl(bg),   brd=bd(col), aln=AL_C)
    for r in range(CR, CR+4):
        ws1.row_dimensions[r].height = 22

    # ── Statut budget objectif ──────────────────────────────────
    SR = CR + 5
    ws1.merge_cells(f"B{SR}:F{SR}")
    csr = ws1.cell(row=SR, column=2)
    if budget_goal > 0:
        if over_budget > 0:
            csr.value = f"Dépassement de budget : {over_budget:,.0f} €  (objectif : {budget_goal:,.0f} €)"
            sc(csr, fnt=fn(bold=True, color=RED, size=11), fll=fl(RED_LT), brd=bd(RED), aln=AL_C)
        else:
            csr.value = f"Budget disponible : {budget_libre:,.0f} €  (objectif : {budget_goal:,.0f} €)"
            sc(csr, fnt=fn(bold=True, color=GREEN, size=11), fll=fl(GREEN_LT), brd=bd(GREEN), aln=AL_C)
    ws1.row_dimensions[SR].height = 22

    # ── Coût par personne ────────────────────────────────────────
    PR = SR + 2
    for i, (lbl, val) in enumerate([
        (f"Coût moyen par invité  ({guest_count} pers.)",  cost_per_guest),
        (f"Coût moyen par adulte  ({adult_count} adultes)", cost_per_adult),
    ]):
        r = PR + i
        c1 = ws1.cell(row=r, column=2, value=lbl)
        c2 = ws1.cell(row=r, column=3, value=val)
        sc(c1, fnt=fn(color=GREY),            fll=fl(PINK_XLT), brd=bd(), aln=AL_L)
        sc(c2, fnt=fn(bold=True, color=PINK_DARK), fll=fl(PINK_XLT), brd=bd(), aln=AL_R, fmt=EUR)
        ws1.row_dimensions[r].height = 18

    # ── Tableau récap par catégorie ──────────────────────────────
    THR = PR + 4
    for j, h in enumerate(["Catégorie", "Prévu", "Payé", "Reste", "% payé"]):
        cc = ws1.cell(row=THR, column=2+j, value=h)
        sc(cc, fnt=fn(bold=True, color=WHITE, size=10), fll=fl(PINK_DARK), brd=bd(PINK), aln=AL_C)
    ws1.row_dimensions[THR].height = 20

    r = THR + 1
    for cat_id in sorted(by_category, key=cat_sort_key):
        items = by_category[cat_id]
        cat_name = categories_by_id.get(cat_id, "Sans catégorie") if cat_id != "__none__" else "Sans catégorie"
        cp = sum(float(i.get("amountTotal", 0) or 0) for i in items)
        cy = sum(float(i.get("amountPaid",  0) or 0) for i in items)
        cr2 = max(cp - cy, 0)
        cpct = cy / cp if cp > 0 else 0
        bg = GREEN_LT if all(i.get("solde") for i in items) else WHITE
        for j, (val, fmt, col, aln) in enumerate([
            (cat_name, None, PINK_DARK, AL_L),
            (cp,  EUR,  BLUE,   AL_R),
            (cy,  EUR,  GREEN,  AL_R),
            (cr2, EUR,  ORANGE if cr2 > 0 else GREY, AL_R),
            (cpct, PCT, GREY,   AL_R),
        ]):
            cc = ws1.cell(row=r, column=2+j, value=val)
            sc(cc, fnt=fn(bold=(j == 0), color=col), fll=fl(bg), brd=bd(), aln=aln, fmt=fmt)
        ws1.row_dimensions[r].height = 18
        r += 1

    tot_pct = total_paye / total_prevu if total_prevu > 0 else 0
    for j, (val, fmt, col, aln) in enumerate([
        ("TOTAL", None, PINK_DARK, AL_L),
        (total_prevu,  EUR, BLUE,   AL_R),
        (total_paye,   EUR, GREEN,  AL_R),
        (total_reste,  EUR, ORANGE if total_reste > 0 else GREY, AL_R),
        (tot_pct, PCT, GREY,   AL_R),
    ]):
        cc = ws1.cell(row=r, column=2+j, value=val)
        sc(cc, fnt=fn(bold=True, color=col, size=11), fll=fl(PINK_LT), brd=bd(PINK), aln=aln, fmt=fmt)
    ws1.row_dimensions[r].height = 22

    # ════════════════════════════════════════════════════════════
    # ONGLET 2 — Dépenses détaillées
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Dépenses détaillées")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [32, 16, 16, 16, 8, 8, 34]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:G1")
    ws2["A1"].value = "Dépenses détaillées"
    sc(ws2["A1"], fnt=fn(bold=True, color=PINK_DARK, size=14),
       fll=fl(PINK_LT), brd=bd(PINK), aln=AL_C)
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:G2")
    c2h = ws2["A2"]
    if updated_at > 0:
        c2h.value = datetime.fromtimestamp(updated_at / 1000)
        c2h.number_format = '"Mis à jour le "DD/MM/YYYY" à "HH:mm'
    sc(c2h, fnt=fn(italic=True, color=GREY, size=9), fll=fl(PINK_XLT), aln=AL_C)
    ws2.row_dimensions[2].height = 14

    # Légende couleurs
    ws2.merge_cells("A4:G4")
    ws2.cell(row=4, column=1).value = "Vert = soldé intégralement   |   Orange = reste à régler   |   Blanc = prévu non encore payé"
    sc(ws2.cell(row=4, column=1), fnt=fn(italic=True, color=GREY, size=9), fll=fl(GREY_LT), aln=AL_C)
    ws2.row_dimensions[4].height = 14

    # En-têtes colonnes
    HDR = 6
    D_HDRS = ["Prestataire", "Montant prévu", "Déjà payé", "Reste à régler", "% payé", "Soldé", "Contact / Notes"]
    for j, h in enumerate(D_HDRS):
        cc = ws2.cell(row=HDR, column=j+1, value=h)
        sc(cc, fnt=fn(bold=True, color=WHITE, size=10), fll=fl(PINK_DARK), brd=bd(PINK), aln=AL_C)
    ws2.row_dimensions[HDR].height = 22

    dr = HDR + 1
    for cat_id in sorted(by_category, key=cat_sort_key):
        items = by_category[cat_id]
        cat_name = categories_by_id.get(cat_id, "Sans catégorie") if cat_id != "__none__" else "Sans catégorie"

        # Entête de catégorie
        ws2.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=7)
        ch = ws2.cell(row=dr, column=1, value=f"  {cat_name.upper()}")
        sc(ch, fnt=fn(bold=True, color=PINK_DARK, size=10), fll=fl(PINK_XLT), brd=bd(PINK_LT), aln=AL_L)
        ws2.row_dimensions[dr].height = 18
        dr += 1

        cat_p = cat_y = 0.0
        for item in items:
            lbl       = str(item.get("label", "")).strip()
            amt_tot   = float(item.get("amountTotal", item.get("amount", 0)) or 0)
            amt_paid  = float(item.get("amountPaid",  0) or 0)
            amt_rest  = max(amt_tot - amt_paid, 0)
            pct       = amt_paid / amt_tot if amt_tot > 0 else 0
            solde     = bool(item.get("solde"))
            sup       = item.get("supplier") or {}
            contact   = str(sup.get("contact", "")).strip()
            notes     = str(sup.get("notes",   "")).strip()
            cn        = " | ".join(x for x in [contact, notes] if x)
            cat_p += amt_tot; cat_y += amt_paid

            row_bg = GREEN_LT if solde else WHITE

            ROW_DEF = [
                (lbl,     None, AL_L, fn(bold=True, color=GREEN if solde else BLACK), row_bg),
                (amt_tot, EUR,  AL_R, fn(bold=True, color=BLUE),  row_bg),
                (amt_paid,EUR,  AL_R, fn(bold=True, color=GREEN), row_bg),
                (amt_rest,EUR,  AL_R, fn(bold=True, color=ORANGE if amt_rest > 0 else GREY),
                                      ORANGE_LT if amt_rest > 0 and not solde else row_bg),
                (pct,     PCT,  AL_C, fn(color=GREEN if pct >= 1 else (ORANGE if pct > 0 else GREY)), row_bg),
                ("Oui" if solde else "", None, AL_C, fn(bold=True, color=GREEN if solde else GREY_LT), row_bg),
                (cn,      None, AL_W, fn(italic=True, color=GREY, size=9), row_bg),
            ]
            for j, (val, fmt, aln, fnt_obj, bg) in enumerate(ROW_DEF):
                cc = ws2.cell(row=dr, column=j+1, value=val)
                sc(cc, fnt=fnt_obj, fll=fl(bg), brd=bd(), aln=aln, fmt=fmt)
            ws2.row_dimensions[dr].height = 20
            dr += 1

        # Sous-total catégorie
        cat_r  = max(cat_p - cat_y, 0)
        cat_pt = cat_y / cat_p if cat_p > 0 else 0
        SUB = [
            (f"Sous-total  {cat_name}", None, AL_L, fn(bold=True, color=PINK_DARK)),
            (cat_p, EUR, AL_R, fn(bold=True, color=BLUE)),
            (cat_y, EUR, AL_R, fn(bold=True, color=GREEN)),
            (cat_r, EUR, AL_R, fn(bold=True, color=ORANGE if cat_r > 0 else GREY)),
            (cat_pt, PCT, AL_R, fn(bold=True, color=GREY)),
            ("", None, AL_C, fn()), ("", None, AL_L, fn()),
        ]
        for j, (val, fmt, aln, fnt_obj) in enumerate(SUB):
            cc = ws2.cell(row=dr, column=j+1, value=val)
            sc(cc, fnt=fnt_obj, fll=fl(PINK_XLT), brd=bd(PINK_LT), aln=aln, fmt=fmt)
        ws2.row_dimensions[dr].height = 18
        dr += 2  # ligne vide entre catégories

    # Total général
    tot_pct2 = total_paye / total_prevu if total_prevu > 0 else 0
    GRAND = [
        ("TOTAL GÉNÉRAL", None, AL_L, fn(bold=True, color=PINK_DARK, size=12)),
        (total_prevu,  EUR, AL_R, fn(bold=True, color=BLUE,   size=12)),
        (total_paye,   EUR, AL_R, fn(bold=True, color=GREEN,  size=12)),
        (total_reste,  EUR, AL_R, fn(bold=True, color=ORANGE if total_reste > 0 else GREY, size=12)),
        (tot_pct2, PCT, AL_R, fn(bold=True, color=GREY, size=12)),
        ("", None, AL_C, fn()), ("", None, AL_L, fn()),
    ]
    for j, (val, fmt, aln, fnt_obj) in enumerate(GRAND):
        cc = ws2.cell(row=dr, column=j+1, value=val)
        sc(cc, fnt=fnt_obj, fll=fl(PINK_LT), brd=bd(PINK), aln=aln, fmt=fmt)
    ws2.row_dimensions[dr].height = 26

    ws2.freeze_panes = f"A{HDR + 1}"
    ws2.auto_filter.ref = f"A{HDR}:G{dr}"

    # ════════════════════════════════════════════════════════════
    # ONGLET 3 — Invités RSVP  (code inchangé)
    # ════════════════════════════════════════════════════════════
    title_fill  = PatternFill(fill_type="solid", fgColor="F6DCE3")
    header_fill = PatternFill(fill_type="solid", fgColor="FCEEF3")
    total_fill  = PatternFill(fill_type="solid", fgColor="F9F3E8")
    title_font  = Font(name="Calibri", size=16, bold=True, color="8A2C4F")
    header_font = Font(name="Calibri", size=11, bold=True, color="7C2044")
    section_font = Font(name="Calibri", size=11, bold=True, color="6B1433")
    value_font  = Font(name="Calibri", size=11, color="3F2330")
    thin_side   = Side(style="thin", color="EBC8D6")
    soft_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centered     = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left",   vertical="center")
    right_aligned = Alignment(horizontal="right", vertical="center")
    currency_format = '#,##0.00 "EUR"'
    percent_format  = "0.00%"

    total_guests   = len(guests)
    adults         = sum(1 for g in guests if g.get("guestCategory") != "child")
    children       = sum(1 for g in guests if g.get("guestCategory") == "child")
    confirmed      = sum(1 for g in guests if g.get("status") == "yes")
    declined       = sum(1 for g in guests if g.get("status") == "no")
    pending        = total_guests - confirmed - declined
    confirm_rate   = confirmed / total_guests if total_guests > 0 else 0
    hebergement_count   = sum(1 for g in guests if g.get("hebergement") and g.get("status") == "yes")
    hebergement_revenue = hebergement_count * 75

    guest_summary_rows = [
        ("Invités au total",              total_guests,        None),
        ("Adultes",                       adults,              None),
        ("Enfants",                       children,            None),
        ("Confirmés",                     confirmed,           None),
        ("Déclinés",                      declined,            None),
        ("En attente",                    pending,             None),
        ("Taux de confirmation",          confirm_rate,        percent_format),
        ("Hébergement — invités (×75 €)", hebergement_count,   None),
        ("Hébergement — revenus",         hebergement_revenue, '#,##0.00 "€"'),
    ]

    guests_sheet = wb.create_sheet("Invités RSVP")
    guests_sheet.sheet_view.showGridLines = False

    guests_sheet.merge_cells("A1:J1")
    t = guests_sheet["A1"]
    t.value = "Liste des invités — Mariage Florence & Antoine"
    t.font = title_font; t.fill = title_fill; t.alignment = centered
    guests_sheet.row_dimensions[1].height = 30

    guests_start_row = 3
    for offset, (label, value, fmt) in enumerate(guest_summary_rows):
        r = guests_start_row + offset
        lc = guests_sheet.cell(row=r, column=1, value=label)
        vc = guests_sheet.cell(row=r, column=2, value=value)
        lc.font = section_font; lc.fill = header_fill
        lc.alignment = left_aligned; lc.border = soft_border
        vc.font = value_font; vc.alignment = right_aligned; vc.border = soft_border
        if fmt:
            vc.number_format = fmt

    guests_table_header_row = guests_start_row + len(guest_summary_rows) + 2
    guest_headers = ["Invité", "Catégorie", "Présence", "Type invitation", "Hébergement", "Info hébergement", "Musique proposée", "Allergies", "Question", "Date RSVP"]
    for col, header in enumerate(guest_headers, start=1):
        cell = guests_sheet.cell(row=guests_table_header_row, column=col, value=header)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = centered; cell.border = soft_border
    guests_sheet.row_dimensions[guests_table_header_row].height = 22

    status_label_map = {"yes": "Confirmé", "no": "Décliné", "pending": "En attente"}
    attendance_label_map = {
        "vin_repas": "Vin d'honneur + repas",
        "vin_only":  "Vin d'honneur",
    }
    status_fill_map = {
        "yes":     PatternFill(fill_type="solid", fgColor="E8F6ED"),
        "no":      PatternFill(fill_type="solid", fgColor="FDECEE"),
        "pending": PatternFill(fill_type="solid", fgColor="FFF9E6"),
    }
    hebergement_fill = PatternFill(fill_type="solid", fgColor="D6EAF8")

    guests_table_start_row = guests_table_header_row + 1
    sorted_guests = sorted(guests, key=lambda g: (
        {"yes": 0, "pending": 1, "no": 2}.get(g.get("status", "pending"), 1),
        str(g.get("name", "")).lower()
    ))
    for index, guest in enumerate(sorted_guests):
        r = guests_table_start_row + index
        name          = str(guest.get("name", "")).strip()
        cat           = guest.get("guestCategory", "adult")
        cat_label     = "Enfant" if cat == "child" else "Adulte"
        status_key    = guest.get("status", "pending") if guest.get("status") in status_label_map else "pending"
        attendance    = normalize_guest_attendance_type(guest.get("attendanceType"))
        hebergement      = bool(guest.get("hebergement", False))
        hebergement_info = bool(guest.get("hebergementInfo", False))
        music            = str(guest.get("musicSuggestion", "")).strip()
        allergies_val    = str(guest.get("allergies", "")).strip()
        question_val     = str(guest.get("otherQuestion", "")).strip()
        submitted_at     = int(guest.get("rsvpSubmittedAt", 0) or 0)
        rsvp_date        = datetime.fromtimestamp(submitted_at / 1000).strftime("%d/%m/%Y") if submitted_at > 0 else ""

        row_values = [
            name,
            cat_label,
            status_label_map[status_key],
            attendance_label_map.get(attendance, "Vin d'honneur + repas"),
            "Oui" if hebergement else "Non",
            "Oui" if hebergement_info else "",
            music,
            allergies_val,
            question_val,
            rsvp_date,
        ]
        aligns = [left_aligned, centered, centered, left_aligned, centered, centered, left_aligned, left_aligned, left_aligned, centered]
        for col, (value, aln) in enumerate(zip(row_values, aligns), start=1):
            cell = guests_sheet.cell(row=r, column=col, value=value)
            cell.font = value_font; cell.border = soft_border; cell.alignment = aln

        guests_sheet.cell(row=r, column=3).fill = status_fill_map[status_key]
        if cat == "child":
            guests_sheet.cell(row=r, column=2).fill = PatternFill(fill_type="solid", fgColor="FFF9E6")
            guests_sheet.cell(row=r, column=2).font = Font(name="Calibri", size=11, color="A04000")
        if hebergement:
            hc = guests_sheet.cell(row=r, column=5)
            hc.fill = hebergement_fill
            hc.font = Font(name="Calibri", size=11, bold=True, color="1F618D")
        if hebergement_info:
            hi = guests_sheet.cell(row=r, column=6)
            hi.fill = PatternFill(fill_type="solid", fgColor="EAF6FF")
            hi.font = Font(name="Calibri", size=11, bold=True, color="1F618D")
        guests_sheet.row_dimensions[r].height = 18

    total_row = guests_table_start_row + len(sorted_guests)
    tl = guests_sheet.cell(row=total_row, column=1, value="TOTAL INVITÉS")
    tv = guests_sheet.cell(row=total_row, column=3, value=total_guests)
    for cell in (tl, tv):
        cell.font = Font(name="Calibri", size=11, bold=True, color="6B1433")
        cell.fill = total_fill; cell.border = soft_border
        cell.alignment = left_aligned if cell.column == 1 else centered

    guests_sheet.column_dimensions["A"].width = 30
    guests_sheet.column_dimensions["B"].width = 11
    guests_sheet.column_dimensions["C"].width = 13
    guests_sheet.column_dimensions["D"].width = 22
    guests_sheet.column_dimensions["E"].width = 13
    guests_sheet.column_dimensions["F"].width = 18
    guests_sheet.column_dimensions["G"].width = 28
    guests_sheet.column_dimensions["H"].width = 24
    guests_sheet.column_dimensions["I"].width = 32
    guests_sheet.column_dimensions["J"].width = 14
    guests_sheet.freeze_panes = f"A{guests_table_start_row}"
    guests_sheet.auto_filter.ref = f"A{guests_table_header_row}:J{total_row}"

    wb.save(BUDGET_EXCEL_FILE)


def load_data_file() -> dict:
    with DATA_LOCK:
        if DATA_FILE.exists():
            try:
                parsed = json.loads(DATA_FILE.read_text(encoding="utf-8"))
                return normalize_data(parsed)
            except (json.JSONDecodeError, OSError):
                pass

        data = create_default_data()
        write_data_file(data)
        return data


def save_data_file(data: dict) -> None:
    normalized = normalize_data(data)
    with DATA_LOCK:
        write_data_file(normalized)
    try:
        write_budget_excel(normalized)
    except Exception as exc:
        print(f"Avertissement Excel : {exc}", flush=True)


def load_admin_password() -> str:
    password = str(os.getenv("MARIAGE_ADMIN_PASSWORD", "")).strip()
    if password:
        print("Mot de passe admin: variable MARIAGE_ADMIN_PASSWORD")
        return password

    print(
        f"Mot de passe admin par défaut utilisé: {DEFAULT_ADMIN_PASSWORD} "
        "(définissez MARIAGE_ADMIN_PASSWORD pour le personnaliser)"
    )
    return DEFAULT_ADMIN_PASSWORD


class PlannerHandler(SimpleHTTPRequestHandler):
    def _send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def end_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "SAMEORIGIN")
        self.send_header("Referrer-Policy", "strict-origin-when-cross-origin")
        super().end_headers()

    def send_error(self, code, message=None, explain=None) -> None:
        if int(code) == 404:
            self._send_html(_404_PAGE, HTTPStatus.NOT_FOUND)
            return
        super().send_error(code, message, explain)

    def _is_admin_authorized(self) -> bool:
        if not ADMIN_PASSWORD:
            return False
        provided = self.headers.get("X-Admin-Key", "")
        return bool(provided) and provided == ADMIN_PASSWORD

    def _check_rsvp_rate_limit(self) -> bool:
        ip = self.client_address[0]
        now = time.time()
        with _RATE_LIMIT_LOCK:
            calls = _RATE_LIMIT[ip]
            calls[:] = [t for t in calls if now - t < 3600]
            if len(calls) >= 10:
                return False
            calls.append(now)
            return True

    def _send_html(self, html_body: str, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = html_body.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _render_rsvp_page(self, token: str, guest_name: str, status: str = "") -> str:
        name = escape(guest_name) if guest_name else "Cher invité"
        if status == "yes":
            title = "Réponse enregistrée: Oui"
            message = "Merci, votre présence est bien confirmée."
        elif status == "no":
            title = "Réponse enregistrée: Non"
            message = "Merci pour votre réponse, elle a bien été enregistrée."
        else:
            title = "Confirmez votre présence"
            message = "Merci de choisir votre réponse."

        yes_url = f"/rsvp?token={quote(token)}&status=yes"
        no_url = f"/rsvp?token={quote(token)}&status=no"

        return f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(title)}</title>
  <style>
    body {{
      margin: 0;
      min-height: 100vh;
      display: grid;
      place-items: center;
      font-family: "Segoe UI", Arial, sans-serif;
      background: linear-gradient(165deg, #fff3f2, #ffdedd);
      color: #4a0b14;
      padding: 18px;
    }}
    .card {{
      width: min(560px, 96vw);
      background: #fff;
      border: 1px solid rgba(224, 10, 38, 0.28);
      border-radius: 18px;
      padding: 22px;
      box-shadow: 0 18px 36px rgba(224, 10, 38, 0.16);
      text-align: center;
    }}
    h1 {{
      margin: 0 0 10px;
      font-size: 1.45rem;
      color: #e00a26;
    }}
    p {{
      margin: 0 0 16px;
      line-height: 1.45;
    }}
    .name {{
      font-weight: 700;
    }}
    .actions {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
    }}
    a {{
      display: inline-block;
      text-decoration: none;
      padding: 10px 12px;
      border-radius: 12px;
      font-weight: 700;
      border: 1px solid rgba(224, 10, 38, 0.6);
    }}
    .yes {{
      background: linear-gradient(145deg, #ffbd1c, #f65d08);
      color: #fff;
    }}
    .no {{
      background: #fff3f2;
      color: #a2182e;
    }}
  </style>
</head>
<body>
  <main class="card">
    <h1>{escape(title)}</h1>
    <p class="name">{name}</p>
    <p>{escape(message)}</p>
    <div class="actions">
      <a class="yes" href="{yes_url}">Oui, je serai présent(e)</a>
      <a class="no" href="{no_url}">Non, je ne pourrai pas venir</a>
    </div>
  </main>
</body>
</html>"""

    def _apply_rsvp_status(self, token: str, status: str) -> tuple[bool, str]:
        if status not in {"yes", "no"}:
            return False, ""

        data = load_data_file()
        for guest in data.get("guests", []):
            if str(guest.get("rsvpToken", "")).strip() == token:
                guest["status"] = status
                data["updatedAt"] = int(time.time() * 1000)
                save_data_file(data)
                return True, str(guest.get("name", "")).strip()
        return False, ""

    def _find_guest_by_token(self, token: str) -> str:
        data = load_data_file()
        for guest in data.get("guests", []):
            if str(guest.get("rsvpToken", "")).strip() == token:
                return str(guest.get("name", "")).strip()
        return ""

    def _is_maintenance_mode(self) -> bool:
        return os.getenv("MARIAGE_MAINTENANCE", "").strip() == "1"

    def _is_access_authorized(self) -> bool:
        if not ACCESS_CODE:
            return True
        expected = _compute_access_token(ACCESS_CODE)
        for part in self.headers.get("Cookie", "").split(";"):
            name, _, value = part.strip().partition("=")
            if name.strip() == ACCESS_COOKIE_NAME and value.strip() == expected:
                return True
        return False

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)

        # ── Access gate (code invités) ────────────────────────────────
        _ACCESS_EXEMPT = {"/access.html", "/api/access"}
        if ACCESS_CODE and path not in _ACCESS_EXEMPT and not self._is_access_authorized():
            if path.startswith("/api/"):
                self._send_json({"ok": False, "error": "access_required"}, HTTPStatus.FORBIDDEN)
            else:
                self.send_response(HTTPStatus.FOUND)
                self.send_header("Location", "/access.html")
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
            return

        # ── Mode maintenance ─────────────────────────────────────────
        if self._is_maintenance_mode() and path != "/api/admin/check" and not self._is_admin_authorized():
            if MAINTENANCE_FILE.exists():
                body = MAINTENANCE_FILE.read_bytes()
                self.send_response(HTTPStatus.SERVICE_UNAVAILABLE)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            else:
                self._send_html(
                    "<!doctype html><html lang='fr'><head><meta charset='UTF-8'><title>Maintenance</title></head>"
                    "<body style='font-family:sans-serif;text-align:center;padding:4rem;'>"
                    "<p style='font-size:3rem;'>💍</p><h1>Nous revenons bientôt</h1>"
                    "<p>Le site est en cours de mise à jour. Revenez dans quelques instants.</p>"
                    "</body></html>",
                    HTTPStatus.SERVICE_UNAVAILABLE,
                )
            return

        if path == "/api/admin/check":
            if not self._is_admin_authorized():
                self._send_json({"ok": False, "error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            self._send_json({"ok": True})
            return

        if path == "/api/data":
            if not self._is_admin_authorized():
                self._send_json({"ok": False, "error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
                return
            self._send_json(load_data_file())
            return

        if path == "/api/stats":
            data = load_data_file()
            guests_list = data.get("guests", [])
            total     = len(guests_list)
            confirmed = sum(1 for g in guests_list if g.get("status") == "yes")
            declined  = sum(1 for g in guests_list if g.get("status") == "no")
            pending   = total - confirmed - declined
            def _party_size(g: dict) -> int:
                gt = str(g.get("groupType", "single"))
                if gt == "single": return 1
                if gt == "couple": return 2
                try: return max(1, int(g.get("partySize", 3)))
                except (TypeError, ValueError): return 3
            total_persons     = sum(_party_size(g) for g in guests_list)
            confirmed_persons = sum(_party_size(g) for g in guests_list if g.get("status") == "yes")
            self._send_json({
                "total": total, "confirmed": confirmed, "declined": declined, "pending": pending,
                "totalPersons": total_persons, "confirmedPersons": confirmed_persons,
            })
            return

        if path == "/api/rsvp/search":
            q = str(query.get("q", [""])[0]).strip()
            if not q or len(q) < 2:
                self._send_json({"ok": True, "results": []})
                return

            def _norm(s: str) -> str:
                s = unicodedata.normalize("NFD", s.lower())
                return "".join(c for c in s if unicodedata.category(c) != "Mn")

            q_words = _norm(q).split()
            data = load_data_file()
            results = []
            for g in data.get("guests", []):
                if g.get("isHost"):
                    continue
                name_norm = _norm(str(g.get("name", "")))
                if all(w in name_norm for w in q_words):
                    results.append({
                        "id": g["id"],
                        "name": g["name"],
                        "rsvpToken": g.get("rsvpToken", ""),
                        "groupType": g.get("groupType", "single"),
                        "attendanceType": g.get("attendanceType", "vin_repas"),
                        "partySize": g.get("partySize", 1),
                        "status": g.get("status", "pending"),
                        "hebergement": g.get("hebergement", False),
                        "hebergementInfo": g.get("hebergementInfo", False),
                        "musicSuggestion": g.get("musicSuggestion", ""),
                        "allergies": g.get("allergies", ""),
                        "otherQuestion": g.get("otherQuestion", ""),
                        "rsvpSubmittedAt": g.get("rsvpSubmittedAt", 0),
                    })
            self._send_json({"ok": True, "results": results})
            return

        # Bloquer l'accès direct aux fichiers sensibles sans auth admin
        BLOCKED_PATHS = {"/data.json", "/data.backup.json", "/budget_mariage.xlsx"}
        if path in BLOCKED_PATHS or path.startswith("/backups/"):
            if not self._is_admin_authorized():
                self.send_error(HTTPStatus.FORBIDDEN, "Accès refusé")
                return

        # Fichiers HTML : servir directement avec no-cache
        if path.endswith(".html") or path == "/":
            fs_path = self.translate_path(self.path)
            import os as _os
            if _os.path.isfile(fs_path):
                with open(fs_path, "rb") as f:
                    content = f.read()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Cache-Control", "no-store, must-revalidate")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return

        super().do_GET()

    def do_POST(self) -> None:
        if self.path == "/api/access":
            content_length = int(self.headers.get("Content-Length", "0"))
            raw_body = self.rfile.read(content_length)
            try:
                payload = json.loads(raw_body.decode("utf-8"))
            except json.JSONDecodeError:
                self.send_error(HTTPStatus.BAD_REQUEST, "Invalid JSON")
                return
            code = str(payload.get("code", "")).strip()
            if ACCESS_CODE and code == ACCESS_CODE:
                token = _compute_access_token(ACCESS_CODE)
                body = json.dumps({"ok": True}, ensure_ascii=False).encode("utf-8")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(body)))
                self.send_header(
                    "Set-Cookie",
                    f"{ACCESS_COOKIE_NAME}={token}; Max-Age=2592000; Path=/; HttpOnly; SameSite=Strict",
                )
                self.end_headers()
                self.wfile.write(body)
            else:
                self._send_json({"ok": False, "error": "wrong_code"}, HTTPStatus.UNAUTHORIZED)
            return

        if self.path == "/api/rsvp/submit":
            if not self._check_rsvp_rate_limit():
                self._send_json({"ok": False, "error": "rate_limit"}, HTTPStatus.TOO_MANY_REQUESTS)
                return
            content_length = int(self.headers.get("Content-Length", "0"))
            raw_body = self.rfile.read(content_length)
            try:
                payload = json.loads(raw_body.decode("utf-8"))
            except json.JSONDecodeError:
                self.send_error(HTTPStatus.BAD_REQUEST, "Invalid JSON")
                return
            guest_id = str(payload.get("id", "")).strip()
            token = str(payload.get("token", "")).strip()
            status = str(payload.get("status", "")).strip()
            hebergement = bool(payload.get("hebergement", False))
            if not guest_id or status not in VALID_GUEST_STATUS:
                self._send_json({"ok": False, "error": "invalid"}, HTTPStatus.BAD_REQUEST)
                return
            data = load_data_file()
            found = False
            for g in data.get("guests", []):
                stored_token = str(g.get("rsvpToken", "")).strip()
                if g["id"] == guest_id and stored_token and stored_token == token:
                    g["status"] = status
                    g["hebergement"] = hebergement
                    g["hebergementInfo"] = bool(payload.get("hebergementInfo", False))
                    g["musicSuggestion"] = str(payload.get("musicSuggestion", "")).strip()
                    g["allergies"] = str(payload.get("allergies", "")).strip()
                    g["otherQuestion"] = str(payload.get("otherQuestion", "")).strip()
                    g["rsvpSubmittedAt"] = int(time.time() * 1000)
                    found = True
                    break
            if not found:
                self._send_json({"ok": False, "error": "not found"}, HTTPStatus.NOT_FOUND)
                return
            data["updatedAt"] = int(time.time() * 1000)
            save_data_file(data)
            # Notification email asynchrone
            submitted_guest = next((g for g in data.get("guests", []) if g["id"] == guest_id), None)
            if submitted_guest:
                send_rsvp_notification(submitted_guest)
            self._send_json({"ok": True})
            return

        if self.path != "/api/data":
            self.send_error(HTTPStatus.NOT_FOUND, "Endpoint not found")
            return

        if not self._is_admin_authorized():
            self._send_json({"ok": False, "error": "unauthorized"}, HTTPStatus.UNAUTHORIZED)
            return

        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length)

        try:
            payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            self.send_error(HTTPStatus.BAD_REQUEST, "Invalid JSON")
            return

        current = load_data_file()
        payload_ts = payload.get("updatedAt", 0)
        current_ts = current.get("updatedAt", 0)
        if isinstance(payload_ts, (int, float)) and isinstance(current_ts, (int, float)):
            if payload_ts > 0 and current_ts > 0 and current_ts > payload_ts:
                self._send_json({"ok": False, "error": "conflict", "serverUpdatedAt": current_ts}, HTTPStatus.CONFLICT)
                return

        try:
            save_data_file(payload)
        except Exception as exc:
            print(f"Erreur sauvegarde data.json : {exc}", flush=True)
            self._send_json({"ok": False, "error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)
            return

        self._send_json({"ok": True})


def main() -> None:
    global ADMIN_PASSWORD, ACCESS_CODE
    ADMIN_PASSWORD = load_admin_password()
    ACCESS_CODE = os.environ.get("MARIAGE_ACCESS_CODE", "").strip()
    if ACCESS_CODE:
        print(f"Protection par code d'accès ACTIVÉE (MARIAGE_ACCESS_CODE défini)", flush=True)
    else:
        print("Protection par code d'accès désactivée (MARIAGE_ACCESS_CODE non défini)", flush=True)
    data = load_data_file()
    write_budget_excel(data)
    start_backup_scheduler()
    handler = partial(PlannerHandler, directory=str(BASE_DIR))
    host = os.environ.get("MARIAGE_HOST", "0.0.0.0")
    port = int(os.environ.get("MARIAGE_PORT", "8000"))
    server = ThreadingHTTPServer((host, port), handler)
    print(f"Serveur actif sur http://{host}:{port}")
    print(f"Fichier de données: {DATA_FILE}")
    print(f"Fichier budget Excel: {BUDGET_EXCEL_FILE}")
    if os.getenv("MARIAGE_MAINTENANCE", "").strip() == "1":
        print("Mode maintenance ACTIVÉ (MARIAGE_MAINTENANCE=1)", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
